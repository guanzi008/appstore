from __future__ import annotations

import os
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.comments import Comment

from appstore.capabilities import build_system_templates, load_capability_cache
from appstore.inspectors import read_package_info
from appstore.models import CapabilityCache, PackageInfo, SystemTemplate

APPS_HEADERS = (
    "app_key",
    "app_name_zh",
    "pkg_name",
    "category_id",
    "website",
    "short_desc_zh",
    "full_desc_zh",
    "icon_path",
    "screenshot_1",
    "screenshot_2",
    "screenshot_3",
    "keywords_zh",
    "app_id_override",
)

RELEASES_HEADERS = (
    "enabled",
    "app_key",
    "release_key",
    "execution_mode",
    "region",
    "note",
)

PACKAGES_BASE_HEADERS = (
    "enabled",
    "app_key",
    "release_key",
    "package_key",
    "file_path",
    "pkg_channel",
    "note",
)

SYSTEM_TEMPLATES_HEADERS = (
    "column_prefix",
    "package_family",
    "system_label",
    "sup_sys_code",
    "baseline_options",
)

SCREENSHOT_COLUMNS = ("screenshot_1", "screenshot_2", "screenshot_3")
_IMAGE_SUFFIXES = {".png", ".jpg", ".jpeg", ".webp"}
_MIN_REAL_IMAGE_BYTES = 1024


@dataclass(frozen=True)
class PreparedNewAppWorkbook:
    output_path: Path
    package_family: str
    package_format: str
    pkg_name: str
    app_key: str
    app_name_zh: str
    release_key: str
    package_paths: tuple[Path, ...]
    selected_system_line_codes: tuple[str, ...]
    missing_fields: tuple[str, ...]
    placeholder_fields: tuple[str, ...]
    auto_detected_assets: dict[str, Any]
    ready_for_upload: bool


def _infer_package_kind(package_path: Path) -> tuple[str, str]:
    suffix = package_path.suffix.lower()
    if suffix == ".deb":
        return "deb", "deb"
    if suffix == ".uab":
        return "linglong", "uab"
    if suffix == ".layer":
        return "linglong", "layer"
    raise ValueError(f"unsupported package format: {package_path.name}")


def _relative_path(path: Path | str, *, from_dir: Path) -> str:
    normalized = Path(path).expanduser().resolve()
    try:
        return os.path.relpath(normalized, from_dir)
    except ValueError:
        return str(normalized)


def _common_parent(paths: tuple[Path, ...]) -> Path:
    if len(paths) == 1:
        return paths[0].parent
    common = os.path.commonpath([str(path.parent.resolve()) for path in paths])
    return Path(common)


def _package_sort_key(info: PackageInfo) -> tuple[int, str, str]:
    arch_order = {"amd64": 0, "arm64": 1, "loong64": 2}
    return (arch_order.get(info.pkg_arch, 99), info.pkg_arch, info.file_path.name)


def _candidate_image(path: Path) -> Path | None:
    if not path.exists() or not path.is_file():
        return None
    if path.suffix.lower() not in _IMAGE_SUFFIXES:
        return None
    if path.stat().st_size < _MIN_REAL_IMAGE_BYTES:
        return None
    return path


def _detect_icon(search_root: Path, *, pkg_name: str) -> Path | None:
    candidates = (
        search_root / "icon.png",
        search_root / f"{pkg_name}.png",
        search_root / "assets" / "icon.png",
        search_root / "assets" / f"{pkg_name}.png",
    )
    for candidate in candidates:
        detected = _candidate_image(candidate)
        if detected is not None:
            return detected
    return None


def _detect_screenshots(search_root: Path) -> tuple[Path, ...]:
    directories = (search_root / "screenshots", search_root / "assets")
    named_candidates = (
        "screenshot_1.png",
        "screenshot_2.png",
        "screenshot_3.png",
        "shot-1.png",
        "shot-2.png",
        "shot-3.png",
    )
    for directory in directories:
        if not directory.exists() or not directory.is_dir():
            continue
        found_named = tuple(
            detected
            for candidate_name in named_candidates
            if (detected := _candidate_image(directory / candidate_name)) is not None
        )
        if found_named:
            return found_named[:3]
        generic = tuple(
            path
            for path in sorted(directory.iterdir())
            if _candidate_image(path) is not None
        )
        if generic:
            return generic[:3]
    return ()


def _load_capability_cache(path: Path | str) -> CapabilityCache:
    cache_path = Path(path)
    if cache_path.is_dir():
        cache_path = cache_path / "latest.json"
    return load_capability_cache(cache_path)


def _template_comment(template: SystemTemplate) -> str:
    if template.baseline_options:
        baseline_text = ", ".join(
            f"{option.baseline_id}:{option.minor_version}" for option in template.baseline_options
        )
    else:
        baseline_text = "无基线要求"
    return (
        f"包类型: {template.package_family}\n"
        f"系统线: {template.system_label} ({template.sup_sys_code})\n"
        f"版本候选: {baseline_text}"
    )


def _package_key(app_key: str, info: PackageInfo, *, used: set[str]) -> str:
    base = f"{app_key}-{info.pkg_arch}"
    candidate = base
    index = 2
    while candidate in used:
        candidate = f"{base}-{index}"
        index += 1
    used.add(candidate)
    return candidate


def prepare_new_app_workbook(
    *,
    packages: list[str] | tuple[str, ...],
    output_path: Path | str,
    capabilities_cache: Path | str,
    app_key: str = "",
    app_name_zh: str = "",
    category_id: int = 1,
    website: str = "",
    short_desc_zh: str = "",
    full_desc_zh: str = "",
    icon_path: str = "",
    screenshot_paths: list[str] | tuple[str, ...] = (),
    keywords_zh: str = "",
    release_key: str = "stable",
    execution_mode: str = "api",
    region: str = "1",
    note: str = "",
    pkg_channel: str = "stable",
    system_line_codes: list[str] | tuple[str, ...] = (),
) -> PreparedNewAppWorkbook:
    if not packages:
        raise ValueError("packages is required")

    output_path = Path(output_path).expanduser().resolve()
    output_path.parent.mkdir(parents=True, exist_ok=True)
    cache = _load_capability_cache(capabilities_cache)

    package_paths = tuple(Path(item).expanduser().resolve() for item in packages)
    package_infos: list[PackageInfo] = []
    for package_path in package_paths:
        package_family, package_format = _infer_package_kind(package_path)
        package_infos.append(read_package_info(package_family, package_format, package_path))
    package_infos.sort(key=_package_sort_key)

    family = package_infos[0].package_family
    package_format = package_infos[0].package_format
    pkg_name = package_infos[0].pkg_name
    pkg_version = package_infos[0].pkg_version
    if any(info.package_family != family for info in package_infos):
        raise ValueError("all packages must belong to the same package family")
    if any(info.pkg_name != pkg_name for info in package_infos):
        raise ValueError("all packages must have the same pkg_name")
    if any(info.pkg_version != pkg_version for info in package_infos):
        raise ValueError("all packages must have the same pkg_version")

    normalized_app_key = app_key.strip() or pkg_name
    normalized_app_name = app_name_zh.strip() or pkg_name
    search_root = _common_parent(package_paths)

    detected_icon = None if icon_path.strip() else _detect_icon(search_root, pkg_name=pkg_name)
    normalized_icon_path = icon_path.strip() or (str(detected_icon) if detected_icon is not None else "")

    normalized_screenshot_inputs = tuple(item.strip() for item in screenshot_paths if str(item).strip())
    detected_screenshots = () if normalized_screenshot_inputs else _detect_screenshots(search_root)
    normalized_screenshots = normalized_screenshot_inputs or tuple(str(path) for path in detected_screenshots)
    screenshot_values = tuple(normalized_screenshots[: len(SCREENSHOT_COLUMNS)])
    if len(screenshot_values) < len(SCREENSHOT_COLUMNS):
        screenshot_values = screenshot_values + ("",) * (len(SCREENSHOT_COLUMNS) - len(screenshot_values))

    templates = tuple(
        template for template in build_system_templates(cache) if template.package_family == family
    )
    selected_codes = tuple(
        code for code in dict.fromkeys(code.strip() for code in system_line_codes if str(code).strip()) if any(template.sup_sys_code == code for template in templates)
    )
    selected_code_set = set(selected_codes)

    workbook = Workbook()
    apps_sheet = workbook.active
    apps_sheet.title = "apps"
    apps_sheet.append(APPS_HEADERS)

    placeholder_website = website.strip() or f"https://example.invalid/{pkg_name}"
    placeholder_short_desc = short_desc_zh.strip() or f"请替换为 {normalized_app_name} 的一句话简介。"
    placeholder_full_desc = full_desc_zh.strip() or f"请替换为 {normalized_app_name} 的详细介绍、主要功能和使用场景。"
    placeholder_keywords = keywords_zh.strip() or normalized_app_name

    apps_sheet.append(
        [
            normalized_app_key,
            normalized_app_name,
            pkg_name,
            category_id,
            placeholder_website,
            placeholder_short_desc,
            placeholder_full_desc,
            _relative_path(normalized_icon_path, from_dir=output_path.parent) if normalized_icon_path else "",
            *(
                _relative_path(value, from_dir=output_path.parent) if value else ""
                for value in screenshot_values
            ),
            placeholder_keywords,
            "",
        ]
    )

    releases_sheet = workbook.create_sheet("releases")
    releases_sheet.append(RELEASES_HEADERS)
    releases_sheet.append([1, normalized_app_key, release_key, execution_mode, region, note])

    system_headers: list[str] = []
    for template in templates:
        system_headers.extend(
            [
                f"{template.column_prefix}__enabled",
                f"{template.column_prefix}__baseline",
                f"{template.column_prefix}__unsupported",
            ]
        )

    packages_sheet = workbook.create_sheet("packages")
    packages_sheet.append(PACKAGES_BASE_HEADERS + tuple(system_headers))
    for index, header in enumerate(system_headers, start=len(PACKAGES_BASE_HEADERS) + 1):
        prefix, field = header.rsplit("__", 1)
        template = next(template for template in templates if template.column_prefix == prefix)
        field_label = {
            "enabled": "是否启用该系统线",
            "baseline": "兼容应用基线，填 baseline_id 或 id:version",
            "unsupported": "不上架版本，填 baseline_id 列表，逗号分隔",
        }[field]
        packages_sheet.cell(row=1, column=index).comment = Comment(
            f"{field_label}\n{_template_comment(template)}",
            "appstore",
        )

    used_package_keys: set[str] = set()
    for info in package_infos:
        package_key = _package_key(normalized_app_key, info, used=used_package_keys)
        row = [
            1,
            normalized_app_key,
            release_key,
            package_key,
            _relative_path(info.file_path, from_dir=output_path.parent),
            pkg_channel,
            f"Auto-filled from real package {info.file_path.name}.",
        ]
        for template in templates:
            if template.sup_sys_code not in selected_code_set:
                row.extend(["", "", ""])
                continue
            baseline_value = template.baseline_options[0].baseline_id if template.baseline_options else ""
            row.extend(["Y", baseline_value, ""])
        packages_sheet.append(row)

    templates_sheet = workbook.create_sheet("system_templates")
    templates_sheet.append(list(SYSTEM_TEMPLATES_HEADERS))
    for template in templates:
        templates_sheet.append(
            [
                template.column_prefix,
                template.package_family,
                template.system_label,
                template.sup_sys_code,
                ", ".join(
                    f"{option.baseline_id}:{option.minor_version}"
                    for option in template.baseline_options
                ),
            ]
        )

    workbook.save(output_path)

    missing_fields: list[str] = []
    if not normalized_icon_path:
        missing_fields.append("icon_path")
    for index, value in enumerate(screenshot_values, start=1):
        if not value:
            missing_fields.append(f"screenshot_{index}")
    if not selected_codes:
        missing_fields.append("system_line_codes")

    placeholder_fields: list[str] = []
    if not app_name_zh.strip():
        placeholder_fields.append("app_name_zh")
    if not website.strip():
        placeholder_fields.append("website")
    if not short_desc_zh.strip():
        placeholder_fields.append("short_desc_zh")
    if not full_desc_zh.strip():
        placeholder_fields.append("full_desc_zh")
    if not keywords_zh.strip():
        placeholder_fields.append("keywords_zh")

    return PreparedNewAppWorkbook(
        output_path=output_path,
        package_family=family,
        package_format=package_format,
        pkg_name=pkg_name,
        app_key=normalized_app_key,
        app_name_zh=normalized_app_name,
        release_key=release_key,
        package_paths=tuple(info.file_path for info in package_infos),
        selected_system_line_codes=selected_codes,
        missing_fields=tuple(missing_fields),
        placeholder_fields=tuple(placeholder_fields),
        auto_detected_assets={
            "icon_path": str(detected_icon) if detected_icon is not None else "",
            "screenshot_paths": [str(path) for path in detected_screenshots],
        },
        ready_for_upload=not missing_fields and not placeholder_fields,
    )
