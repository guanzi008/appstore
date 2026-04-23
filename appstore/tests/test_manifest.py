from __future__ import annotations

import unittest
from pathlib import Path
from tempfile import TemporaryDirectory

from openpyxl import Workbook

from appstore.manifest import ManifestError, _normalize_enabled, load_manifest


class ManifestLoaderTests(unittest.TestCase):
    def _build_workbook(
        self,
        root: Path,
        *,
        duplicate_app: bool = False,
        legacy_releases: bool = False,
        include_packages: bool = True,
        include_targets: bool = True,
        duplicate_release: bool = False,
        duplicate_package: bool = False,
        duplicate_target: bool = False,
        unknown_target_package: bool = False,
        trimmed_values: bool = False,
        target_rows: int = 1,
    ) -> Path:
        workbook = Workbook()

        apps = workbook.active
        apps.title = "apps"
        apps.append(
            [
                "app_key",
                "app_name_zh",
                "pkg_name",
                "category_id",
                "website",
                "short_desc_zh",
                "full_desc_zh",
                "icon_path",
                "screenshot_1",
                "screenshot_2",
                "screenshot_3",
                "keywords_zh",
                "app_id_override",
            ]
        )
        app_row = [
            "demo",
            "演示应用",
            "demo",
            1,
            "https://example.invalid/demo",
            "简短介绍",
            "详细介绍",
            "assets/icon.png",
            "assets/shot-1.png",
            "assets/shot-2.png",
            "assets/shot-3.png",
            "标签,打印",
            "",
        ]
        if trimmed_values:
            app_row = [
                " demo ",
                " 演示应用 ",
                " demo ",
                1,
                " https://example.invalid/demo ",
                " 简短介绍 ",
                " 详细介绍 ",
                " assets/icon.png ",
                " assets/shot-1.png ",
                " assets/shot-2.png ",
                " assets/shot-3.png ",
                " 标签,打印 ",
                " override ",
            ]
        apps.append(app_row)
        if duplicate_app:
            apps.append(
                [
                    "demo",
                    "重复应用",
                    "demo",
                    1,
                    "https://example.invalid/dup",
                    "重复",
                    "重复",
                    "assets/icon.png",
                    "assets/shot-1.png",
                    "assets/shot-2.png",
                    "assets/shot-3.png",
                    "",
                    "",
                ]
            )

        releases = workbook.create_sheet("releases")
        if legacy_releases:
            releases.append(["enabled", "app_key", "deb_path", "system_platform", "arch", "region", "baseline", "note"])
        else:
            releases.append(["enabled", "app_key", "release_key", "release_name", "region", "note"])
            release_row = ["Y", "demo", "stable", "稳定版", "cn", "release note"]
            if trimmed_values:
                release_row = [" yes ", " demo ", " stable ", " 稳定版 ", " cn ", " release note "]
            releases.append(release_row)
            if duplicate_release:
                releases.append(["Y", "demo", "stable", "稳定版-duplicate", "cn", "duplicate release"])
            releases.append(["N", "demo", "ignored", "忽略版", "cn", "disabled release"])

        if include_packages:
            packages = workbook.create_sheet("packages")
            packages.append(
                [
                    "enabled",
                    "app_key",
                    "release_key",
                    "package_key",
                    "package_family",
                    "package_format",
                    "file_path",
                    "declared_arch",
                    "pkg_channel",
                    "note",
                ]
            )
            package_row = [
                "yes",
                "demo",
                "stable",
                "pkg-amd64",
                "demo",
                "deb",
                "packages/demo_1.0.0_amd64.deb",
                "amd64",
                "stable",
                "package note",
            ]
            if trimmed_values:
                package_row = [
                    " yes ",
                    " demo ",
                    " stable ",
                    " pkg-amd64 ",
                    " demo ",
                    " deb ",
                    " packages/demo_1.0.0_amd64.deb ",
                    " amd64 ",
                    " stable ",
                    " package note ",
                ]
            packages.append(package_row)
            if duplicate_package:
                packages.append(
                    [
                        "yes",
                        "demo",
                        "stable",
                        "pkg-amd64",
                        "demo",
                        "deb",
                        "packages/demo_1.0.0_amd64.deb",
                        "amd64",
                        "stable",
                        "duplicate package",
                    ]
                )
            packages.append(
                [
                    "no",
                    "demo",
                    "stable",
                    "pkg-disabled",
                    "demo",
                    "deb",
                    "packages/ignored.deb",
                    "amd64",
                    "stable",
                    "disabled package",
                ]
            )

        if include_targets:
            targets = workbook.create_sheet("targets")
            targets.append(
                [
                    "enabled",
                    "app_key",
                    "release_key",
                    "package_key",
                    "sup_sys_code",
                    "baseline_id",
                    "unsupport_baseline_ids",
                    "target_note",
                ]
            )
            target_row = ["true", "demo", "stable", "pkg-amd64", "Deepin_23", "23.0", "20, 21, ", "target note"]
            if trimmed_values:
                target_row = [" yes ", " demo ", " stable ", " pkg-amd64 ", " Deepin_23 ", " 23.0 ", " 20, 21, ", " target note "]
            if unknown_target_package:
                target_row[3] = "missing-package"
            if target_rows > 0:
                targets.append(target_row)
            if duplicate_target:
                targets.append(["true", "demo", "stable", "pkg-amd64", "Deepin_23", "23.0", "20, 21, ", "duplicate target"])
            targets.append(["false", "demo", "stable", "pkg-disabled", "Deepin_23", "23.0", "", "disabled target"])

        workbook_path = root / "manifest.xlsx"
        workbook.save(workbook_path)
        return workbook_path

    def test_load_manifest_normalizes_workbook_records(self) -> None:
        with TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            (root / "assets").mkdir()
            (root / "packages").mkdir()
            for name in ("icon.png", "shot-1.png", "shot-2.png", "shot-3.png"):
                (root / "assets" / name).write_bytes(b"asset")
            (root / "packages" / "demo_1.0.0_amd64.deb").write_bytes(b"deb-bytes")

            manifest = load_manifest(self._build_workbook(root))

            self.assertEqual(set(manifest.apps.keys()), {"demo"})
            self.assertEqual(set(manifest.releases.keys()), {("demo", "stable")})
            self.assertEqual(set(manifest.packages.keys()), {("demo", "stable")})
            self.assertEqual(set(manifest.targets.keys()), {("demo", "stable", "pkg-amd64")})

            app = manifest.apps["demo"]
            self.assertEqual(app.icon_path, root / "assets" / "icon.png")
            self.assertEqual(
                app.screenshot_paths,
                (
                    root / "assets" / "shot-1.png",
                    root / "assets" / "shot-2.png",
                    root / "assets" / "shot-3.png",
                ),
            )

            release = manifest.releases[("demo", "stable")]
            self.assertEqual(release.release_key, "stable")
            self.assertEqual(release.release_name, "稳定版")
            self.assertEqual(release.region, "cn")
            self.assertEqual(release.note, "release note")

            package = manifest.packages[("demo", "stable")]
            self.assertEqual(len(package), 1)
            self.assertEqual(package[0].file_path, root / "packages" / "demo_1.0.0_amd64.deb")
            self.assertEqual(package[0].declared_arch, "amd64")
            self.assertEqual(package[0].pkg_channel, "stable")
            self.assertEqual(package[0].note, "package note")

            target = manifest.targets[("demo", "stable", "pkg-amd64")]
            self.assertEqual(len(target), 1)
            self.assertEqual(target[0].sup_sys_code, "Deepin_23")
            self.assertEqual(target[0].baseline_id, "23.0")
            self.assertEqual(target[0].unsupport_baseline_ids, ("20", "21"))
            self.assertEqual(target[0].target_note, "target note")

    def test_load_manifest_rejects_missing_required_sheet(self) -> None:
        with TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            (root / "assets").mkdir()
            (root / "packages").mkdir()
            for name in ("icon.png", "shot-1.png", "shot-2.png", "shot-3.png"):
                (root / "assets" / name).write_bytes(b"asset")
            (root / "packages" / "demo_1.0.0_amd64.deb").write_bytes(b"deb-bytes")

            workbook_path = self._build_workbook(root, include_targets=False)

            with self.assertRaisesRegex(ManifestError, "missing targets sheet or package system template columns"):
                load_manifest(workbook_path)

    def test_load_manifest_rejects_legacy_release_schema(self) -> None:
        with TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            (root / "assets").mkdir()
            (root / "packages").mkdir()
            for name in ("icon.png", "shot-1.png", "shot-2.png", "shot-3.png"):
                (root / "assets" / name).write_bytes(b"asset")
            (root / "packages" / "demo_1.0.0_amd64.deb").write_bytes(b"deb-bytes")

            workbook_path = self._build_workbook(root, legacy_releases=True)

            with self.assertRaisesRegex(
                ManifestError,
                "legacy releases schema unsupported; use releases/packages sheets",
            ):
                load_manifest(workbook_path)

    def test_load_manifest_rejects_duplicate_release_key(self) -> None:
        with TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            (root / "assets").mkdir()
            (root / "packages").mkdir()
            for name in ("icon.png", "shot-1.png", "shot-2.png", "shot-3.png"):
                (root / "assets" / name).write_bytes(b"asset")
            (root / "packages" / "demo_1.0.0_amd64.deb").write_bytes(b"deb-bytes")

            workbook_path = self._build_workbook(root, duplicate_release=True)

            with self.assertRaisesRegex(ManifestError, "duplicate release_key"):
                load_manifest(workbook_path)

    def test_load_manifest_rejects_duplicate_package_key(self) -> None:
        with TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            (root / "assets").mkdir()
            (root / "packages").mkdir()
            for name in ("icon.png", "shot-1.png", "shot-2.png", "shot-3.png"):
                (root / "assets" / name).write_bytes(b"asset")
            (root / "packages" / "demo_1.0.0_amd64.deb").write_bytes(b"deb-bytes")

            workbook_path = self._build_workbook(root, duplicate_package=True)

            with self.assertRaisesRegex(ManifestError, "duplicate package_key"):
                load_manifest(workbook_path)

    def test_load_manifest_rejects_duplicate_target_key(self) -> None:
        with TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            (root / "assets").mkdir()
            (root / "packages").mkdir()
            for name in ("icon.png", "shot-1.png", "shot-2.png", "shot-3.png"):
                (root / "assets" / name).write_bytes(b"asset")
            (root / "packages" / "demo_1.0.0_amd64.deb").write_bytes(b"deb-bytes")

            workbook_path = self._build_workbook(root, duplicate_target=True)

            with self.assertRaisesRegex(ManifestError, "duplicate target_key"):
                load_manifest(workbook_path)

    def test_load_manifest_rejects_unknown_package_reference_from_target(self) -> None:
        with TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            (root / "assets").mkdir()
            (root / "packages").mkdir()
            for name in ("icon.png", "shot-1.png", "shot-2.png", "shot-3.png"):
                (root / "assets" / name).write_bytes(b"asset")
            (root / "packages" / "demo_1.0.0_amd64.deb").write_bytes(b"deb-bytes")

            workbook_path = self._build_workbook(root, unknown_target_package=True)

            with self.assertRaisesRegex(ManifestError, "unknown package_key in targets row"):
                load_manifest(workbook_path)

    def test_load_manifest_rejects_package_without_targets(self) -> None:
        with TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            (root / "assets").mkdir()
            (root / "packages").mkdir()
            for name in ("icon.png", "shot-1.png", "shot-2.png", "shot-3.png"):
                (root / "assets" / name).write_bytes(b"asset")
            (root / "packages" / "demo_1.0.0_amd64.deb").write_bytes(b"deb-bytes")

            workbook_path = self._build_workbook(root, target_rows=0)

            with self.assertRaisesRegex(ManifestError, "package has no targets"):
                load_manifest(workbook_path)

    def test_load_manifest_skips_disabled_rows(self) -> None:
        with TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            (root / "assets").mkdir()
            (root / "packages").mkdir()
            for name in ("icon.png", "shot-1.png", "shot-2.png", "shot-3.png"):
                (root / "assets" / name).write_bytes(b"asset")
            (root / "packages" / "demo_1.0.0_amd64.deb").write_bytes(b"deb-bytes")

            manifest = load_manifest(self._build_workbook(root))
            self.assertEqual(set(manifest.releases.keys()), {("demo", "stable")})
            self.assertEqual(set(manifest.packages.keys()), {("demo", "stable")})
            self.assertEqual(set(manifest.targets.keys()), {("demo", "stable", "pkg-amd64")})

    def test_load_manifest_derives_targets_from_package_template_columns(self) -> None:
        with TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            (root / "assets").mkdir()
            (root / "packages").mkdir()
            for name in ("icon.png", "shot-1.png", "shot-2.png", "shot-3.png"):
                (root / "assets" / name).write_bytes(b"asset")
            (root / "packages" / "demo_1.0.0_amd64.deb").write_bytes(b"deb-bytes")

            workbook = Workbook()
            apps = workbook.active
            apps.title = "apps"
            apps.append(
                [
                    "app_key",
                    "app_name_zh",
                    "pkg_name",
                    "category_id",
                    "website",
                    "short_desc_zh",
                    "full_desc_zh",
                    "icon_path",
                    "screenshot_1",
                    "screenshot_2",
                    "screenshot_3",
                ]
            )
            apps.append(
                [
                    "demo",
                    "演示应用",
                    "demo",
                    1,
                    "https://example.invalid/demo",
                    "简短介绍",
                    "详细介绍",
                    "assets/icon.png",
                    "assets/shot-1.png",
                    "assets/shot-2.png",
                    "assets/shot-3.png",
                ]
            )

            releases = workbook.create_sheet("releases")
            releases.append(["enabled", "app_key", "release_key", "execution_mode", "region", "note"])
            releases.append(["Y", "demo", "stable", "auto", "cn", "release note"])

            packages = workbook.create_sheet("packages")
            packages.append(
                [
                    "enabled",
                    "app_key",
                    "release_key",
                    "package_key",
                    "file_path",
                    "pkg_channel",
                    "note",
                    "sys__deb__11__enabled",
                    "sys__deb__11__baseline",
                    "sys__deb__11__unsupported",
                ]
            )
            packages.append(
                [
                    "Y",
                    "demo",
                    "stable",
                    "pkg-amd64",
                    "packages/demo_1.0.0_amd64.deb",
                    "stable",
                    "package note",
                    "Y",
                    "2300",
                    "2301,2302",
                ]
            )

            workbook_path = root / "template-manifest.xlsx"
            workbook.save(workbook_path)

            manifest = load_manifest(workbook_path)
            target = manifest.targets[("demo", "stable", "pkg-amd64")][0]
            self.assertEqual(target.sup_sys_code, "11")
            self.assertEqual(target.baseline_id, "2300")
            self.assertEqual(target.unsupport_baseline_ids, ("2301", "2302"))

    def test_load_manifest_trims_string_values(self) -> None:
        with TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            (root / "assets").mkdir()
            (root / "packages").mkdir()
            for name in ("icon.png", "shot-1.png", "shot-2.png", "shot-3.png"):
                (root / "assets" / name).write_bytes(b"asset")
            (root / "packages" / "demo_1.0.0_amd64.deb").write_bytes(b"deb-bytes")

            manifest = load_manifest(self._build_workbook(root, trimmed_values=True))

            self.assertEqual(set(manifest.apps.keys()), {"demo"})
            self.assertEqual(set(manifest.releases.keys()), {("demo", "stable")})
            self.assertEqual(set(manifest.packages.keys()), {("demo", "stable")})
            self.assertEqual(set(manifest.targets.keys()), {("demo", "stable", "pkg-amd64")})
            self.assertEqual(manifest.apps["demo"].app_id_override, "override")
            self.assertEqual(manifest.releases[("demo", "stable")].release_name, "稳定版")
            self.assertEqual(manifest.packages[("demo", "stable")][0].package_key, "pkg-amd64")
            self.assertEqual(manifest.targets[("demo", "stable", "pkg-amd64")][0].baseline_id, "23.0")

    def test_load_manifest_rejects_duplicate_app_keys(self) -> None:
        with TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            (root / "assets").mkdir()
            (root / "packages").mkdir()
            for name in ("icon.png", "shot-1.png", "shot-2.png", "shot-3.png"):
                (root / "assets" / name).write_bytes(b"asset")
            (root / "packages" / "demo_1.0.0_amd64.deb").write_bytes(b"deb-bytes")

            workbook_path = self._build_workbook(root, duplicate_app=True)

            with self.assertRaisesRegex(ManifestError, "duplicate app_key"):
                load_manifest(workbook_path)

    def test_normalize_enabled_recognizes_only_known_true_markers(self) -> None:
        self.assertFalse(_normalize_enabled("tru"))
        self.assertFalse(_normalize_enabled("maybe"))
        self.assertTrue(_normalize_enabled(" yes "))
        self.assertTrue(_normalize_enabled("yes"))
        self.assertTrue(_normalize_enabled("Y"))


if __name__ == "__main__":
    unittest.main()
