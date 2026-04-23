import subprocess
import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from appstore.capabilities import CapabilityCache, write_capability_cache
from appstore.deb import read_deb_package_info
from appstore.inspectors import inspect_package
from appstore.manifest import load_manifest
from appstore.models import BaselineOption, SystemLine
from appstore.submission import validate_release_group


APPS_HEADERS = [
    "app_key",
    "app_name_zh",
    "pkg_name",
    "category_id",
    "website",
    "short_desc_zh",
    "full_desc_zh",
    "icon_path",
    "screenshot_1",
    "screenshot_2",
    "screenshot_3",
    "keywords_zh",
    "app_id_override",
]

RELEASES_HEADERS = [
    "enabled",
    "app_key",
    "release_key",
    "execution_mode",
    "region",
    "note",
]

PACKAGES_BASE_HEADERS = [
    "enabled",
    "app_key",
    "release_key",
    "package_key",
    "file_path",
    "pkg_channel",
    "note",
]

SYSTEM_TEMPLATES_HEADERS = [
    "column_prefix",
    "package_family",
    "system_label",
    "sup_sys_code",
    "baseline_options",
]

README_REQUIRED_SNIPPETS = [
    "sync-capabilities",
    "validate",
    "upload",
    "generate-template",
    "`packages`",
    "`system_templates`",
]


class GenerateTemplateTests(unittest.TestCase):
    def test_generate_template_creates_example_workbook_and_supporting_files(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            output_path = Path(temp_dir) / "template.xlsx"
            cache_dir = Path(temp_dir) / "cache"
            script_path = Path(__file__).resolve().parents[1] / "examples" / "generate_template.py"
            cache = CapabilityCache(
                generated_at="2026-04-23T12:00:00+08:00",
                deb_system_lines={
                    "11": SystemLine(code="11", label="communityV23", family="deb"),
                    "21": SystemLine(code="21", label="communityV25", family="deb"),
                },
                linglong_system_lines={},
                baseline_options={
                    "deb:11": (BaselineOption(system_line_code="11", baseline_id="2300", minor_version="23.0.0"),),
                    "deb:21": (BaselineOption(system_line_code="21", baseline_id="2500", minor_version="25.0.0"),),
                },
            )
            write_capability_cache(cache_dir, cache)

            subprocess.run(
                [sys.executable, str(script_path), str(output_path), str(cache_dir)],
                check=True,
            )

            workbook = load_workbook(output_path)
            self.assertEqual(workbook.sheetnames, ["apps", "releases", "packages", "system_templates"])
            self.assertEqual(
                [cell.value for cell in workbook["apps"][1]],
                APPS_HEADERS,
            )
            self.assertEqual(
                [cell.value for cell in workbook["releases"][1]],
                RELEASES_HEADERS,
            )
            self.assertEqual(workbook["releases"]["D2"].value, "auto")
            self.assertEqual(
                [cell.value for cell in workbook["packages"][1]][: len(PACKAGES_BASE_HEADERS)],
                PACKAGES_BASE_HEADERS,
            )
            self.assertEqual(
                [cell.value for cell in workbook["packages"][1]][len(PACKAGES_BASE_HEADERS):],
                [
                    "sys__deb__11__enabled",
                    "sys__deb__11__baseline",
                    "sys__deb__11__unsupported",
                    "sys__deb__21__enabled",
                    "sys__deb__21__baseline",
                    "sys__deb__21__unsupported",
                ],
            )
            self.assertEqual(workbook["packages"].max_row, 4)
            self.assertEqual(
                [cell.value for cell in workbook["packages"][2]][3:13],
                [
                    "labelnova-amd64",
                    "packages/labelnova_1.0.4-1_amd64.deb",
                    "stable",
                    "Real amd64 package from OBS Deepin_23.",
                    "Y",
                    "2300",
                    None,
                    None,
                    None,
                    None,
                ],
            )
            self.assertEqual(
                [cell.value for cell in workbook["packages"][3]][3:13],
                [
                    "labelnova-arm64",
                    "packages/labelnova_1.0.4-1_arm64.deb",
                    "stable",
                    "Real arm64 package from OBS Deepin_23.",
                    "Y",
                    "2300",
                    None,
                    None,
                    None,
                    None,
                ],
            )
            self.assertEqual(
                [cell.value for cell in workbook["packages"][4]][3:13],
                [
                    "labelnova-loong64",
                    "packages/labelnova_1.0.4-1_loong64.deb",
                    "stable",
                    "Real loong64 package from OBS update flow.",
                    None,
                    None,
                    None,
                    "Y",
                    "2500",
                    None,
                ],
            )
            self.assertEqual(
                [cell.value for cell in workbook["system_templates"][1]],
                SYSTEM_TEMPLATES_HEADERS,
            )
            self.assertEqual(workbook["system_templates"].max_row, 3)
            self.assertEqual(
                [cell.value for cell in workbook["system_templates"][2]],
                ["sys__deb__11", "deb", "communityV23", "11", "2300:23.0.0"],
            )
            self.assertEqual(
                [cell.value for cell in workbook["system_templates"][3]],
                ["sys__deb__21", "deb", "communityV25", "21", "2500:25.0.0"],
            )

            manifest = load_manifest(output_path)
            self.assertEqual(set(manifest.apps.keys()), {"labelnova"})
            self.assertEqual(set(manifest.releases.keys()), {("labelnova", "stable")})
            self.assertEqual(set(manifest.packages.keys()), {("labelnova", "stable")})
            self.assertEqual(
                set(manifest.targets.keys()),
                {
                    ("labelnova", "stable", "labelnova-amd64"),
                    ("labelnova", "stable", "labelnova-arm64"),
                    ("labelnova", "stable", "labelnova-loong64"),
                },
            )
            self.assertEqual(manifest.targets[("labelnova", "stable", "labelnova-amd64")][0].sup_sys_code, "11")
            self.assertEqual(manifest.targets[("labelnova", "stable", "labelnova-arm64")][0].baseline_id, "2300")
            self.assertEqual(manifest.targets[("labelnova", "stable", "labelnova-loong64")][0].sup_sys_code, "21")
            self.assertEqual(manifest.targets[("labelnova", "stable", "labelnova-loong64")][0].baseline_id, "2500")

            release = manifest.releases[("labelnova", "stable")]
            packages = manifest.packages[("labelnova", "stable")]
            validated = validate_release_group(
                app=manifest.apps["labelnova"],
                release=release,
                packages=packages,
                targets_by_package={
                    package.package_key: manifest.targets[(package.app_key, package.release_key, package.package_key)]
                    for package in packages
                },
                inspected_by_package={package.package_key: inspect_package(package) for package in packages},
                capability_cache=cache,
            )
            self.assertEqual(validated.package_family, "deb")
            self.assertEqual(len(validated.packages), 3)

            self.assertTrue((output_path.parent / "assets" / "icon.png").exists())
            expected_packages = {
                "labelnova_1.0.4-1_amd64.deb": "amd64",
                "labelnova_1.0.4-1_arm64.deb": "arm64",
                "labelnova_1.0.4-1_loong64.deb": "loong64",
            }
            for package_name, expected_arch in expected_packages.items():
                package_path = output_path.parent / "packages" / package_name
                self.assertTrue(package_path.exists())
                package_info = read_deb_package_info(package_path)
                self.assertEqual(package_info.pkg_name, "labelnova")
                self.assertEqual(package_info.pkg_version, "1.0.4-1")
                self.assertEqual(package_info.pkg_arch, expected_arch)

    def test_readme_mentions_grouped_cli_and_normalized_sheets(self) -> None:
        readme = (Path(__file__).resolve().parents[1] / "README.md").read_text(encoding="utf-8")
        for snippet in README_REQUIRED_SNIPPETS:
            self.assertIn(snippet, readme)
