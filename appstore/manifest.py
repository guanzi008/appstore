from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from appstore.models import AppRecord, LoadedManifest, PackageRecord, ReleaseRecord, TargetRecord

REQUIRED_APP_COLUMNS = (
    "app_key",
    "app_name_zh",
    "pkg_name",
    "category_id",
    "website",
    "short_desc_zh",
    "full_desc_zh",
    "icon_path",
    "screenshot_1",
    "screenshot_2",
    "screenshot_3",
)

REQUIRED_RELEASE_COLUMNS = (
    "enabled",
    "app_key",
    "release_key",
)

REQUIRED_PACKAGE_COLUMNS = (
    "enabled",
    "app_key",
    "release_key",
    "package_key",
    "file_path",
)

REQUIRED_TARGET_COLUMNS = (
    "enabled",
    "app_key",
    "release_key",
    "package_key",
    "sup_sys_code",
)

LEGACY_RELEASE_COLUMNS = {"deb_path", "system_platform", "arch"}


class ManifestError(ValueError):
    pass


def _is_blank(value: Any) -> bool:
    return value is None or (isinstance(value, str) and not value.strip())


def _normalize_cell(value: Any) -> Any:
    if isinstance(value, str):
        return value.strip()
    return value


def _text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _require(row: dict[str, Any], column: str, row_id: int) -> Any:
    value = row.get(column)
    if _is_blank(value):
        raise ManifestError(f"missing {column} in row {row_id}")
    return value


def _normalize_enabled(value: Any) -> bool:
    if value is None:
        return False
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value != 0
    if isinstance(value, str):
        normalized = value.strip().lower()
        if normalized in {"1", "y", "yes", "true", "on", "enabled"}:
            return True
    return False


def _resolve_path(workbook_path: Path, raw_value: Any) -> Path:
    path = Path(raw_value)
    if path.is_absolute():
        return path
    return workbook_path.parent / path


def _sheet_rows(sheet):
    rows = sheet.iter_rows(values_only=True)
    try:
        headers = next(rows)
    except StopIteration:
        return

    header_names = [_normalize_cell(header) if header is not None else "" for header in headers]
    for row_id, values in enumerate(rows, start=2):
        normalized_values = tuple(_normalize_cell(value) for value in values)
        if all(value in (None, "") for value in normalized_values):
            continue
        yield row_id, {
            header_names[index]: value
            for index, value in enumerate(normalized_values)
            if index < len(header_names) and header_names[index]
        }


def _sheet_headers(sheet) -> set[str]:
    header_rows = sheet.iter_rows(min_row=1, max_row=1, values_only=True)
    try:
        headers = next(header_rows)
    except StopIteration:
        return set()
    return {str(header).strip() for header in headers if header is not None and str(header).strip()}


def _csv_tokens(value: Any) -> tuple[str, ...]:
    if value is None:
        return ()
    tokens = []
    for token in str(value).split(","):
        normalized = token.strip()
        if normalized:
            tokens.append(normalized)
    return tuple(tokens)


def _require_sheet(workbook, sheet_name: str):
    if sheet_name not in workbook.sheetnames:
        raise ManifestError(f"missing {sheet_name} sheet")
    return workbook[sheet_name]


def _optional_sheet(workbook, sheet_name: str):
    if sheet_name not in workbook.sheetnames:
        return None
    return workbook[sheet_name]


def _append_grouped_record(
    collection: dict[tuple[str, ...], tuple[Any, ...]],
    key: tuple[str, ...],
    record: Any,
    *,
    duplicate_error: str,
    is_duplicate,
) -> None:
    existing = collection.get(key, ())
    if any(is_duplicate(existing_record, record) for existing_record in existing):
        raise ManifestError(duplicate_error)
    collection[key] = existing + (record,)


def _legacy_template_headers(package_headers: set[str]) -> tuple[str, ...]:
    return tuple(sorted(header for header in package_headers if header.startswith("tpl__")))


def _parse_legacy_template_header(header: str) -> tuple[str, str, str]:
    parts = header.split("__")
    if len(parts) != 4 or parts[0] != "tpl":
        raise ManifestError(f"invalid system template column: {header}")
    package_family, sup_sys_code, baseline_id = parts[1], parts[2], parts[3]
    if package_family not in {"deb", "linglong"}:
        raise ManifestError(f"invalid system template column family: {header}")
    return package_family, sup_sys_code, "" if baseline_id == "none" else baseline_id


def _system_field_headers(package_headers: set[str]) -> dict[tuple[str, str], dict[str, str]]:
    grouped: dict[tuple[str, str], dict[str, str]] = {}
    for header in sorted(package_headers):
        if not header.startswith("sys__"):
            continue
        parts = header.split("__")
        if len(parts) != 4 or parts[0] != "sys":
            raise ManifestError(f"invalid system template column: {header}")
        package_family, sup_sys_code, field_name = parts[1], parts[2], parts[3]
        if package_family not in {"deb", "linglong"}:
            raise ManifestError(f"invalid system template column family: {header}")
        if field_name not in {"enabled", "baseline", "unsupported"}:
            raise ManifestError(f"invalid system template column field: {header}")
        grouped.setdefault((package_family, sup_sys_code), {})[field_name] = header
    return grouped


def _normalize_baseline_value(value: Any) -> str:
    normalized = _text(value)
    if not normalized:
        return ""
    return normalized.split(":", 1)[0].strip()


def _infer_package_kind(file_path: Path) -> tuple[str, str]:
    suffix = file_path.suffix.lower()
    if suffix == ".deb":
        return "deb", "deb"
    if suffix == ".uab":
        return "linglong", "uab"
    if suffix == ".layer":
        return "linglong", "layer"
    raise ManifestError(f"unsupported package format for file: {file_path.name}")


def load_manifest(workbook_path: Path | str) -> LoadedManifest:
    workbook_path = Path(workbook_path)
    workbook = load_workbook(workbook_path)

    apps_sheet = _require_sheet(workbook, "apps")
    releases_sheet = _require_sheet(workbook, "releases")
    packages_sheet = _require_sheet(workbook, "packages")
    targets_sheet = _optional_sheet(workbook, "targets")

    app_headers = _sheet_headers(apps_sheet)
    release_headers = _sheet_headers(releases_sheet)
    package_headers = _sheet_headers(packages_sheet)
    app_rows = list(_sheet_rows(apps_sheet))
    release_rows = list(_sheet_rows(releases_sheet))
    package_rows = list(_sheet_rows(packages_sheet))
    target_headers = _sheet_headers(targets_sheet) if targets_sheet is not None else set()
    target_rows = list(_sheet_rows(targets_sheet)) if targets_sheet is not None else []
    legacy_template_headers = _legacy_template_headers(package_headers)
    system_field_headers = _system_field_headers(package_headers)

    missing_app_columns = [column for column in REQUIRED_APP_COLUMNS if column not in app_headers]
    if missing_app_columns:
        raise ManifestError(f"missing required app columns: {', '.join(missing_app_columns)}")

    if LEGACY_RELEASE_COLUMNS.intersection(release_headers):
        raise ManifestError("legacy releases schema unsupported; use releases/packages sheets")

    missing_release_columns = [column for column in REQUIRED_RELEASE_COLUMNS if column not in release_headers]
    if missing_release_columns:
        raise ManifestError(f"missing required release columns: {', '.join(missing_release_columns)}")

    missing_package_columns = [column for column in REQUIRED_PACKAGE_COLUMNS if column not in package_headers]
    if missing_package_columns:
        raise ManifestError(f"missing required package columns: {', '.join(missing_package_columns)}")

    if targets_sheet is not None:
        missing_target_columns = [column for column in REQUIRED_TARGET_COLUMNS if column not in target_headers]
        if missing_target_columns:
            raise ManifestError(f"missing required target columns: {', '.join(missing_target_columns)}")

    if targets_sheet is not None and (legacy_template_headers or system_field_headers):
        raise ManifestError("mixed target schema unsupported; use targets sheet or package template columns")
    if targets_sheet is None and not legacy_template_headers and not system_field_headers:
        raise ManifestError("missing targets sheet or package system template columns")

    apps: dict[str, AppRecord] = {}
    for row_id, row in app_rows:
        app_key = str(_require(row, "app_key", row_id))
        if app_key in apps:
            raise ManifestError(f"duplicate app_key: {app_key}")

        screenshot_paths = tuple(
            _resolve_path(workbook_path, _require(row, column, row_id))
            for column in ("screenshot_1", "screenshot_2", "screenshot_3")
        )
        apps[app_key] = AppRecord(
            app_key=app_key,
            app_name_zh=str(_require(row, "app_name_zh", row_id)),
            pkg_name=str(_require(row, "pkg_name", row_id)),
            category_id=int(_require(row, "category_id", row_id)),
            website=str(_require(row, "website", row_id)),
            short_desc_zh=str(_require(row, "short_desc_zh", row_id)),
            full_desc_zh=str(_require(row, "full_desc_zh", row_id)),
            icon_path=_resolve_path(workbook_path, _require(row, "icon_path", row_id)),
            screenshot_paths=screenshot_paths,
            keywords_zh="" if row.get("keywords_zh") in (None,) else str(row.get("keywords_zh", "")),
            app_id_override="" if row.get("app_id_override") in (None,) else str(row.get("app_id_override", "")),
        )

    releases: dict[tuple[str, str], ReleaseRecord] = {}
    for row_id, row in release_rows:
        if not _normalize_enabled(row.get("enabled")):
            continue

        app_key = str(_require(row, "app_key", row_id))
        if app_key not in apps:
            raise ManifestError(f"unknown app_key in releases row {row_id}: {app_key}")

        release_key = _text(_require(row, "release_key", row_id))
        release_record = ReleaseRecord(
            row_id=row_id,
            app_key=app_key,
            release_key=release_key,
            release_name=_text(row.get("release_name")) or release_key,
            execution_mode=_text(row.get("execution_mode")),
            region=_text(row.get("region")),
            note=_text(row.get("note")),
        )
        if (app_key, release_key) in releases:
            raise ManifestError(f"duplicate release_key: {app_key}/{release_key}")
        releases[(app_key, release_key)] = release_record

    packages: dict[tuple[str, str], tuple[PackageRecord, ...]] = {}
    for row_id, row in package_rows:
        if not _normalize_enabled(row.get("enabled")):
            continue

        app_key = _text(_require(row, "app_key", row_id))
        release_key = _text(_require(row, "release_key", row_id))
        release_id = (app_key, release_key)
        if release_id not in releases:
            raise ManifestError(f"unknown release_key in packages row {row_id}: {app_key}/{release_key}")

        file_path = _resolve_path(workbook_path, _require(row, "file_path", row_id))
        package_family = _text(row.get("package_family"))
        package_format = _text(row.get("package_format"))
        if not package_family or not package_format:
            package_family, package_format = _infer_package_kind(file_path)

        package_record = PackageRecord(
            row_id=row_id,
            app_key=app_key,
            release_key=release_key,
            package_key=_text(_require(row, "package_key", row_id)),
            package_family=package_family,
            package_format=package_format,
            file_path=file_path,
            declared_arch=_text(row.get("declared_arch")),
            pkg_channel=_text(row.get("pkg_channel")),
            note=_text(row.get("note")),
        )
        _append_grouped_record(
            packages,
            release_id,
            package_record,
            duplicate_error=f"duplicate package_key: {app_key}/{release_key}/{package_record.package_key}",
            is_duplicate=lambda existing_record, candidate: existing_record.package_key == candidate.package_key,
        )

    targets: dict[tuple[str, str, str], tuple[TargetRecord, ...]] = {}
    if targets_sheet is not None:
        for row_id, row in target_rows:
            if not _normalize_enabled(row.get("enabled")):
                continue

            app_key = _text(_require(row, "app_key", row_id))
            release_key = _text(_require(row, "release_key", row_id))
            package_key = _text(_require(row, "package_key", row_id))
            release_id = (app_key, release_key)
            package_id = (app_key, release_key)
            if release_id not in releases:
                raise ManifestError(f"unknown release_key in targets row {row_id}: {app_key}/{release_key}")
            if package_id not in packages or not any(package.package_key == package_key for package in packages[package_id]):
                raise ManifestError(f"unknown package_key in targets row {row_id}: {app_key}/{release_key}/{package_key}")

            target_record = TargetRecord(
                row_id=row_id,
                app_key=app_key,
                release_key=release_key,
                package_key=package_key,
                sup_sys_code=_text(_require(row, "sup_sys_code", row_id)),
                baseline_id=_text(row.get("baseline_id")),
                unsupport_baseline_ids=_csv_tokens(row.get("unsupport_baseline_ids")),
                target_note=_text(row.get("target_note")),
            )
            target_group_key = (app_key, release_key, package_key)
            _append_grouped_record(
                targets,
                target_group_key,
                target_record,
                duplicate_error=f"duplicate target_key: {app_key}/{release_key}/{package_key}",
                is_duplicate=lambda existing_record, candidate: (
                    existing_record.sup_sys_code == candidate.sup_sys_code
                    and existing_record.baseline_id == candidate.baseline_id
                    and existing_record.unsupport_baseline_ids == candidate.unsupport_baseline_ids
                ),
            )
    else:
        for release_id, package_group in packages.items():
            for package_record in package_group:
                target_group_key = (package_record.app_key, package_record.release_key, package_record.package_key)
                matching_row = next(
                    (
                        row
                        for row_id, row in package_rows
                        if _normalize_enabled(row.get("enabled"))
                        and _text(row.get("app_key")) == package_record.app_key
                        and _text(row.get("release_key")) == package_record.release_key
                        and _text(row.get("package_key")) == package_record.package_key
                    ),
                    None,
                )
                if matching_row is None:
                    continue
                for header in legacy_template_headers:
                    if not _normalize_enabled(matching_row.get(header)):
                        continue
                    template_family, sup_sys_code, baseline_id = _parse_legacy_template_header(header)
                    if template_family != package_record.package_family:
                        raise ManifestError(
                            f"package template family mismatch for {package_record.package_key}: {header}"
                        )
                    target_record = TargetRecord(
                        row_id=package_record.row_id,
                        app_key=package_record.app_key,
                        release_key=package_record.release_key,
                        package_key=package_record.package_key,
                        sup_sys_code=sup_sys_code,
                        baseline_id=baseline_id,
                    )
                    _append_grouped_record(
                        targets,
                        target_group_key,
                        target_record,
                        duplicate_error=f"duplicate target_key: {package_record.app_key}/{package_record.release_key}/{package_record.package_key}",
                        is_duplicate=lambda existing_record, candidate: (
                            existing_record.sup_sys_code == candidate.sup_sys_code
                            and existing_record.baseline_id == candidate.baseline_id
                            and existing_record.unsupport_baseline_ids == candidate.unsupport_baseline_ids
                        ),
                    )
                for (template_family, sup_sys_code), field_headers in system_field_headers.items():
                    has_enabled = _normalize_enabled(matching_row.get(field_headers.get("enabled", "")))
                    baseline_id = _normalize_baseline_value(matching_row.get(field_headers.get("baseline", "")))
                    unsupported_ids = _csv_tokens(matching_row.get(field_headers.get("unsupported", "")))
                    if not has_enabled and not baseline_id and not unsupported_ids:
                        continue
                    if template_family != package_record.package_family:
                        raise ManifestError(
                            f"package template family mismatch for {package_record.package_key}: {template_family}/{sup_sys_code}"
                        )
                    target_record = TargetRecord(
                        row_id=package_record.row_id,
                        app_key=package_record.app_key,
                        release_key=package_record.release_key,
                        package_key=package_record.package_key,
                        sup_sys_code=sup_sys_code,
                        baseline_id=baseline_id,
                        unsupport_baseline_ids=unsupported_ids,
                    )
                    _append_grouped_record(
                        targets,
                        target_group_key,
                        target_record,
                        duplicate_error=f"duplicate target_key: {package_record.app_key}/{package_record.release_key}/{package_record.package_key}",
                        is_duplicate=lambda existing_record, candidate: (
                            existing_record.sup_sys_code == candidate.sup_sys_code
                            and existing_record.baseline_id == candidate.baseline_id
                            and existing_record.unsupport_baseline_ids == candidate.unsupport_baseline_ids
                        ),
                    )

    for (app_key, release_key), package_group in packages.items():
        for package_record in package_group:
            if (app_key, release_key, package_record.package_key) not in targets:
                raise ManifestError(
                    f"package has no targets: {app_key}/{release_key}/{package_record.package_key}"
                )

    return LoadedManifest(workbook_path=workbook_path, apps=apps, releases=releases, packages=packages, targets=targets)
