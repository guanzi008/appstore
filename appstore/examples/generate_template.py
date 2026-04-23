from __future__ import annotations

import base64
import io
import sys
import tarfile
import time
from pathlib import Path

if __package__ in (None, ""):
    sys.path.insert(0, str(Path(__file__).resolve().parents[2]))

from openpyxl import Workbook
from openpyxl.comments import Comment

from appstore.capabilities import build_system_templates, load_capability_cache
from appstore.models import BaselineOption, CapabilityCache, SystemLine, SystemTemplate

DEFAULT_OUTPUT = Path("appstore/examples/template.xlsx")
DEFAULT_CAPABILITY_CACHE = Path("appstore/cache/capabilities")

APPS_HEADERS = (
    "app_key",
    "app_name_zh",
    "pkg_name",
    "category_id",
    "website",
    "short_desc_zh",
    "full_desc_zh",
    "icon_path",
    "screenshot_1",
    "screenshot_2",
    "screenshot_3",
    "keywords_zh",
    "app_id_override",
)

RELEASES_HEADERS = (
    "enabled",
    "app_key",
    "release_key",
    "execution_mode",
    "region",
    "note",
)

PACKAGES_BASE_HEADERS = (
    "enabled",
    "app_key",
    "release_key",
    "package_key",
    "file_path",
    "pkg_channel",
    "note",
)

SAMPLE_APP_ROW = (
    "labelnova",
    "LabelNova",
    "labelnova",
    1,
    "https://example.com/labelnova",
    "标签打印工具",
    "用于设计并打印标签的示例应用。",
    "assets/icon.png",
    "assets/shot-1.png",
    "assets/shot-2.png",
    "assets/shot-3.png",
    "标签,打印,示例",
    "",
)

SAMPLE_RELEASE_ROW = (
    1,
    "labelnova",
    "stable",
    "auto",
    "1",
    "Example release row for dry-run validation.",
)

SAMPLE_PACKAGE_ROWS = (
    {
        "enabled": 1,
        "app_key": "labelnova",
        "release_key": "stable",
        "package_key": "labelnova-amd64",
        "file_path": "packages/labelnova_1.0.4-1_amd64.deb",
        "pkg_channel": "stable",
        "note": "Real amd64 package from OBS Deepin_23.",
    },
    {
        "enabled": 1,
        "app_key": "labelnova",
        "release_key": "stable",
        "package_key": "labelnova-arm64",
        "file_path": "packages/labelnova_1.0.4-1_arm64.deb",
        "pkg_channel": "stable",
        "note": "Real arm64 package from OBS Deepin_23.",
    },
    {
        "enabled": 1,
        "app_key": "labelnova",
        "release_key": "stable",
        "package_key": "labelnova-loong64",
        "file_path": "packages/labelnova_1.0.4-1_loong64.deb",
        "pkg_channel": "stable",
        "note": "Real loong64 package from OBS update flow.",
    },
)

SAMPLE_PACKAGE_FILES = (
    ("labelnova_1.0.4-1_amd64.deb", "amd64"),
    ("labelnova_1.0.4-1_arm64.deb", "arm64"),
    ("labelnova_1.0.4-1_loong64.deb", "loong64"),
)

_PLACEHOLDER_PNG = base64.b64decode(
    "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAQAAAC1HAwCAAAAC0lEQVR42mP8/x8AAusB9Wn8f1kAAAAASUVORK5CYII="
)


def _bundled_capability_cache() -> CapabilityCache:
    return CapabilityCache(
        generated_at="2026-04-23T00:00:00+08:00",
        deb_system_lines={
            "11": SystemLine(code="11", label="communityV23", family="deb"),
            "21": SystemLine(code="21", label="communityV25", family="deb"),
        },
        linglong_system_lines={
            "11": SystemLine(code="11", label="communityV23", family="linglong"),
            "21": SystemLine(code="21", label="communityV25", family="linglong"),
        },
        baseline_options={
            "deb:11": (BaselineOption(system_line_code="11", baseline_id="2300", minor_version="23.0.0"),),
            "deb:21": (BaselineOption(system_line_code="21", baseline_id="2500", minor_version="25.0.0"),),
            "linglong:11": (BaselineOption(system_line_code="11", baseline_id="2300", minor_version="23.0.0"),),
            "linglong:21": (BaselineOption(system_line_code="21", baseline_id="2500", minor_version="25.0.0"),),
        },
    )


def _tar_gz_bytes(files: dict[str, bytes]) -> bytes:
    buffer = io.BytesIO()
    with tarfile.open(fileobj=buffer, mode="w:gz") as archive:
        for name, data in files.items():
            info = tarfile.TarInfo(name=name)
            info.size = len(data)
            info.mode = 0o644
            info.mtime = 0
            archive.addfile(info, io.BytesIO(data))
    return buffer.getvalue()


def _ar_member(name: str, data: bytes) -> bytes:
    timestamp = int(time.time())
    header = (
        f"{name:<16}"
        f"{timestamp:<12}"
        f"{0:<6}"
        f"{0:<6}"
        f"{0o100644:<8}"
        f"{len(data):<10}`\n"
    ).encode("ascii")
    if len(header) != 60:
        raise ValueError("invalid ar header size")
    padding = b"\n" if len(data) % 2 else b""
    return header + data + padding


def _build_sample_deb(arch: str) -> bytes:
    control = "\n".join(
        [
            "Package: labelnova",
            "Version: 1.0.4-1",
            f"Architecture: {arch}",
            "Maintainer: Example Maintainer <maintainer@example.com>",
            "Description: Minimal valid example package for appstore dry-run",
            "",
        ]
    ).encode("utf-8")
    readme = b"LabelNova example payload\n"
    control_archive = _tar_gz_bytes({"./control": control})
    data_archive = _tar_gz_bytes({"./usr/share/doc/labelnova/README": readme})
    return b"!<arch>\n" + b"".join(
        [
            _ar_member("debian-binary", b"2.0\n"),
            _ar_member("control.tar.gz", control_archive),
            _ar_member("data.tar.gz", data_archive),
        ]
    )

_PLACEHOLDER_DEBS = {
    arch: _build_sample_deb(arch)
    for _filename, arch in SAMPLE_PACKAGE_FILES
}


def _write_placeholder_files(output_path: Path) -> None:
    assets_dir = output_path.parent / "assets"
    packages_dir = output_path.parent / "packages"
    bundled_packages_dir = Path(__file__).resolve().parent / "packages"
    assets_dir.mkdir(parents=True, exist_ok=True)
    packages_dir.mkdir(parents=True, exist_ok=True)

    for name in ("icon.png", "shot-1.png", "shot-2.png", "shot-3.png"):
        (assets_dir / name).write_bytes(_PLACEHOLDER_PNG)

    for package_name, arch in SAMPLE_PACKAGE_FILES:
        destination = packages_dir / package_name
        bundled_package = bundled_packages_dir / package_name
        if bundled_package.exists():
            if bundled_package.resolve() != destination.resolve():
                destination.write_bytes(bundled_package.read_bytes())
            continue
        destination.write_bytes(_PLACEHOLDER_DEBS[arch])


def _load_template_capability_cache(path: Path | str | None) -> CapabilityCache:
    if path is not None:
        cache_path = Path(path)
        if cache_path.exists():
            return load_capability_cache(cache_path)
        if cache_path == DEFAULT_CAPABILITY_CACHE:
            return _bundled_capability_cache()
        raise FileNotFoundError(f"capability cache not found: {cache_path}")
    if DEFAULT_CAPABILITY_CACHE.exists():
        return load_capability_cache(DEFAULT_CAPABILITY_CACHE)
    return _bundled_capability_cache()


def _template_comment(template: SystemTemplate) -> str:
    if template.baseline_options:
        baseline_text = ", ".join(
            f"{option.baseline_id}:{option.minor_version}" for option in template.baseline_options
        )
    else:
        baseline_text = "无基线要求"
    return (
        f"包类型: {template.package_family}\n"
        f"系统线: {template.system_label} ({template.sup_sys_code})\n"
        f"版本候选: {baseline_text}"
    )


def _select_template_keys(templates: tuple[SystemTemplate, ...]) -> dict[str, str]:
    template_by_prefix = {template.column_prefix: template for template in templates}
    deb_templates = [template for template in templates if template.package_family == "deb"]
    if not deb_templates:
        return {}

    v23 = template_by_prefix.get("sys__deb__11", deb_templates[0])
    v25 = template_by_prefix.get("sys__deb__21", deb_templates[-1])
    return {
        "labelnova-amd64": v23.column_prefix,
        "labelnova-arm64": v23.column_prefix,
        "labelnova-loong64": v25.column_prefix,
    }


def _build_workbook(capability_cache: CapabilityCache) -> Workbook:
    workbook = Workbook()
    templates = build_system_templates(capability_cache)
    system_headers: list[str] = []
    for template in templates:
        system_headers.extend(
            [
                f"{template.column_prefix}__enabled",
                f"{template.column_prefix}__baseline",
                f"{template.column_prefix}__unsupported",
            ]
        )
    template_headers = tuple(system_headers)
    template_selection = _select_template_keys(templates)

    apps_sheet = workbook.active
    apps_sheet.title = "apps"
    apps_sheet.append(APPS_HEADERS)
    apps_sheet.append(SAMPLE_APP_ROW)

    releases_sheet = workbook.create_sheet("releases")
    releases_sheet.append(RELEASES_HEADERS)
    releases_sheet.append(SAMPLE_RELEASE_ROW)

    packages_sheet = workbook.create_sheet("packages")
    packages_sheet.append(PACKAGES_BASE_HEADERS + template_headers)
    for index, header in enumerate(template_headers, start=len(PACKAGES_BASE_HEADERS) + 1):
        prefix, field = header.rsplit("__", 1)
        template = next(template for template in templates if template.column_prefix == prefix)
        field_label = {
            "enabled": "是否启用该系统线",
            "baseline": "兼容应用基线，填 baseline_id 或 id:version",
            "unsupported": "不上架版本，填 baseline_id 列表，逗号分隔",
        }[field]
        packages_sheet.cell(row=1, column=index).comment = Comment(
            f"{field_label}\n{_template_comment(template)}",
            "appstore",
        )

    for row in SAMPLE_PACKAGE_ROWS:
        package_values = [row.get(header, "") for header in PACKAGES_BASE_HEADERS]
        selected_prefix = template_selection.get(row["package_key"], "")
        for header in template_headers:
            prefix, field = header.rsplit("__", 1)
            selected_template = next((template for template in templates if template.column_prefix == prefix), None)
            if prefix != selected_prefix:
                package_values.append("")
                continue
            if field == "enabled":
                package_values.append("Y")
            elif field == "baseline":
                package_values.append(selected_template.baseline_options[0].baseline_id if selected_template and selected_template.baseline_options else "")
            else:
                package_values.append("")
        packages_sheet.append(package_values)

    templates_sheet = workbook.create_sheet("system_templates")
    templates_sheet.append(
        [
            "column_prefix",
            "package_family",
            "system_label",
            "sup_sys_code",
            "baseline_options",
        ]
    )
    for template in templates:
        templates_sheet.append(
            [
                template.column_prefix,
                template.package_family,
                template.system_label,
                template.sup_sys_code,
                ", ".join(f"{option.baseline_id}:{option.minor_version}" for option in template.baseline_options),
            ]
        )
    return workbook


def generate_template(
    output_path: Path | str = DEFAULT_OUTPUT,
    *,
    capability_cache_path: Path | str | None = None,
) -> Path:
    destination = Path(output_path)
    destination.parent.mkdir(parents=True, exist_ok=True)

    workbook = _build_workbook(_load_template_capability_cache(capability_cache_path))
    workbook.save(destination)
    _write_placeholder_files(destination)
    return destination


def main(argv: list[str] | None = None) -> int:
    args = list(sys.argv[1:] if argv is None else argv)
    output_path = Path(args[0]) if args else DEFAULT_OUTPUT
    capability_cache_path = Path(args[1]) if len(args) > 1 else None
    generate_template(output_path, capability_cache_path=capability_cache_path)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
